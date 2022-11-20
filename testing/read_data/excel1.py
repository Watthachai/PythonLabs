import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheetNames=wb.sheetnames

for name in sheetNames:
   print(name)

sheet = wb["Sheet1"]
max_row = sheet.max_row
max_col = sheet.max_column
print(max_col,max_row)

for i in range(1, max_row+1, 2):
    print(i, sheet.cell(row=i, column=2).value,end='')
    print(i, sheet.cell(row=i, column=3).value)
